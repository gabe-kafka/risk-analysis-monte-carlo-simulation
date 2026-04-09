#!/usr/bin/env python3
"""
Monte Carlo Schedule Risk Simulation
Matches Primavera Risk Analysis (OPRA) methodology.

Uncertainty is modeled solely through three-point (triangular) duration
distributions per activity. Risk register findings are reflected in the
distribution parameters (wider tails on risk-exposed activities).

Inputs:
  - Schedule CSV:  activity network with durations and predecessors
  - Risk params CSV: optimistic/most-likely/pessimistic factors per activity
  - OR an Excel workbook containing both (created with --init)

Outputs (written to outputs/ and back into the workbook):
  - Histogram of project finish dates
  - S-curve (CDF) with P50, P80, P90 markers
  - Tornado chart — activity sensitivity (Spearman rank correlation)
  - Results sheet — summary statistics and full percentile table
  - Sensitivity sheet — all activities ranked by correlation

Usage:
  python3 scripts/monte_carlo.py --schedule data/activities.csv --risk-params data/risk-params.csv --init
  python3 scripts/monte_carlo.py --workbook data/monte-carlo.xlsx
  python3 scripts/monte_carlo.py --workbook data/monte-carlo.xlsx -n 50000 --seed 123
"""

from __future__ import annotations

import argparse
import csv
import random
from collections import deque
from dataclasses import dataclass, field
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Defaults ─────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

DEFAULT_N = 10_000
DEFAULT_SEED = 42

# ── Style ────────────────────────────────────────────────────────────────
BG = "#FFFFFF"
TEXT = "#1A1A1A"
MUTED = "#6B6B6B"
ACCENT = "#3B82F6"
CRIT = "#DC2626"
GREEN = "#16A34A"

WBS_PALETTE = [
    "#264653", "#2A9D8F", "#457B9D", "#8D99AE", "#E76F51",
    "#F4A261", "#6A994E", "#BC6C25", "#7B2CBF", "#3B82F6",
    "#06B6D4", "#DC2626", "#D97706", "#16A34A", "#A855F7",
]
_wbs_color_cache: dict[str, str] = {}


def wbs_color(wbs: str) -> str:
    """Auto-assign colors to WBS codes in order encountered."""
    if wbs not in _wbs_color_cache:
        _wbs_color_cache[wbs] = WBS_PALETTE[len(_wbs_color_cache) % len(WBS_PALETTE)]
    return _wbs_color_cache[wbs]

# ── Excel style constants ────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="264653", end_color="264653", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D4D4D4"),
    right=Side(style="thin", color="D4D4D4"),
    top=Side(style="thin", color="D4D4D4"),
    bottom=Side(style="thin", color="D4D4D4"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = CENTER


# ── Data structures ──────────────────────────────────────────────────────
@dataclass
class Activity:
    wbs: str
    wbs_name: str
    activity_id: str
    activity_name: str
    duration_days: int
    predecessors: list[str]
    successors: list[str] = field(default_factory=list)


@dataclass
class RiskParams:
    optimistic_factor: float
    most_likely_factor: float
    pessimistic_factor: float


# ── Load schedule from CSV (always) ─────────────────────────────────────
def load_schedule(path: Path) -> tuple[dict[str, Activity], list[str]]:
    activities: dict[str, Activity] = {}
    order: list[str] = []
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            preds = [p.strip() for p in row["predecessors"].split(",") if p.strip()]
            act = Activity(
                wbs=row["wbs"], wbs_name=row["wbs_name"],
                activity_id=row["activity_id"], activity_name=row["activity_name"],
                duration_days=int(row["duration_days"]), predecessors=preds,
            )
            activities[act.activity_id] = act
            order.append(act.activity_id)
    for act in activities.values():
        for pred_id in act.predecessors:
            activities[pred_id].successors.append(act.activity_id)
    return activities, order


# ── Excel workbook: read assumptions ─────────────────────────────────────
def load_from_workbook(path: Path) -> dict[str, RiskParams]:
    wb = load_workbook(path, data_only=True)

    ws1 = wb["Duration Uncertainty"]
    risk_params: dict[str, RiskParams] = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        aid = str(row[0]).strip()
        risk_params[aid] = RiskParams(
            optimistic_factor=float(row[5]),
            most_likely_factor=float(row[6]),
            pessimistic_factor=float(row[7]),
        )

    wb.close()
    return risk_params


# ── Excel workbook: write results back ───────────────────────────────────
def write_results_to_workbook(
    path: Path,
    deterministic: float,
    finishes: np.ndarray,
    correlations: list[tuple[str, str, float, str]],
    seed: int = DEFAULT_SEED,
) -> None:
    wb = load_workbook(path)

    # ── Results sheet ────────────────────────────────────────────────
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results", 2)
    ws.sheet_properties.tabColor = "16A34A"

    # Summary block
    summary_headers = ["Metric", "Value"]
    for c, h in enumerate(summary_headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 2)

    p_vals = [5, 10, 20, 25, 30, 40, 50, 60, 70, 75, 80, 85, 90, 95]
    det_conf = np.searchsorted(np.sort(finishes), deterministic) / len(finishes) * 100

    rows = [
        ("Iterations", f"{len(finishes):,}"),
        ("Seed", str(seed)),
        ("", ""),
        ("Deterministic Finish (days)", f"{deterministic:.0f}"),
        ("Deterministic Confidence", f"{det_conf:.0f}%"),
        ("", ""),
        ("Mean (days)", f"{np.mean(finishes):.0f}"),
        ("Std Dev (days)", f"{np.std(finishes):.1f}"),
        ("Min (days)", f"{np.min(finishes):.0f}"),
        ("Max (days)", f"{np.max(finishes):.0f}"),
        ("", ""),
    ]
    for p in p_vals:
        rows.append((f"P{p}", f"{np.percentile(finishes, p):.0f}"))

    for i, (metric, value) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2).border = THIN_BORDER
        if metric == "":
            continue
        ws.cell(row=i, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16

    # ── Sensitivity sheet ────────────────────────────────────────────
    if "Sensitivity" in wb.sheetnames:
        del wb["Sensitivity"]
    ws2 = wb.create_sheet("Sensitivity", 3)
    ws2.sheet_properties.tabColor = "F4A261"

    sens_headers = ["Rank", "Activity ID", "Activity Name", "WBS",
                    "Spearman ρ", "Interpretation"]
    for c, h in enumerate(sens_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header_row(ws2, 1, len(sens_headers))

    sorted_corr = sorted(correlations, key=lambda x: abs(x[2]), reverse=True)
    for i, (aid, name, corr, wbs) in enumerate(sorted_corr, 2):
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=aid)
        ws2.cell(row=i, column=3, value=name)
        ws2.cell(row=i, column=4, value=wbs)
        ws2.cell(row=i, column=5, value=round(corr, 4))
        ws2.cell(row=i, column=5).number_format = "0.000"
        # interpretation
        ac = abs(corr)
        if ac >= 0.4:
            interp = "Strong driver"
        elif ac >= 0.2:
            interp = "Moderate driver"
        elif ac >= 0.1:
            interp = "Minor driver"
        else:
            interp = "Negligible"
        ws2.cell(row=i, column=6, value=interp)

    end_sens = 1 + len(sorted_corr)
    style_data_rows(ws2, 2, end_sens, len(sens_headers))

    widths_sens = [6, 12, 40, 6, 12, 18]
    for c, w in enumerate(widths_sens, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


# ── CPM engine ───────────────────────────────────────────────────────────
def topological_order(activities: dict[str, Activity], csv_order: list[str]) -> list[str]:
    idx = {aid: i for i, aid in enumerate(csv_order)}
    indegree = {aid: len(a.predecessors) for aid, a in activities.items()}
    ready = deque(sorted([aid for aid, d in indegree.items() if d == 0], key=idx.get))
    ordered: list[str] = []
    while ready:
        cur = ready.popleft()
        ordered.append(cur)
        for succ in sorted(activities[cur].successors, key=idx.get):
            indegree[succ] -= 1
            if indegree[succ] == 0:
                ready.append(succ)
    return ordered


def forward_pass(topo: list[str], activities: dict[str, Activity],
                 durations: dict[str, float]) -> dict[str, float]:
    start: dict[str, float] = {}
    finish: dict[str, float] = {}
    for aid in topo:
        act = activities[aid]
        if act.predecessors:
            start[aid] = max(finish[p] for p in act.predecessors)
        else:
            start[aid] = 0.0
        finish[aid] = start[aid] + durations[aid]
    return finish


# ── Simulation ───────────────────────────────────────────────────────────
def run_simulation(
    activities: dict[str, Activity], csv_order: list[str],
    risk_params: dict[str, RiskParams],
    n_iter: int, seed: int,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    rng = random.Random(seed)
    topo = topological_order(activities, csv_order)
    project_finishes = np.zeros(n_iter)
    activity_durations: dict[str, np.ndarray] = {aid: np.zeros(n_iter) for aid in csv_order}

    for i in range(n_iter):
        sampled: dict[str, float] = {}
        for aid in csv_order:
            base = activities[aid].duration_days
            rp = risk_params[aid]
            lo = base * rp.optimistic_factor
            mode = base * rp.most_likely_factor
            hi = base * rp.pessimistic_factor
            sampled[aid] = rng.triangular(lo, hi, mode)

        finish = forward_pass(topo, activities, sampled)
        project_finishes[i] = max(finish.values())
        for aid in csv_order:
            activity_durations[aid][i] = sampled[aid]

    return project_finishes, activity_durations


# ── Plotting ─────────────────────────────────────────────────────────────
def plot_histogram(finishes: np.ndarray, deterministic: float, out_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(12, 6), facecolor=BG)
    ax.set_facecolor(BG)
    ax.hist(finishes, bins=80, color=ACCENT, alpha=0.75, edgecolor="#2563EB", linewidth=0.3)

    p50, p80, p90 = (np.percentile(finishes, p) for p in [50, 80, 90])
    ax.axvline(deterministic, color=GREEN, linewidth=2, linestyle="--", label=f"Deterministic: {deterministic:.0f}d")
    ax.axvline(p50, color="#F59E0B", linewidth=2, linestyle="-", label=f"P50: {p50:.0f}d")
    ax.axvline(p80, color="#F97316", linewidth=2, linestyle="-", label=f"P80: {p80:.0f}d")
    ax.axvline(p90, color=CRIT, linewidth=2, linestyle="-", label=f"P90: {p90:.0f}d")

    ax.set_xlabel("Project Duration (Working Days)", fontsize=11, fontfamily="monospace", color=TEXT)
    ax.set_ylabel("Frequency", fontsize=11, fontfamily="monospace", color=TEXT)
    ax.set_title("Monte Carlo Schedule Simulation\n"
                 f"N = {len(finishes):,} iterations  |  Three-point triangular duration distributions",
                 fontsize=13, fontfamily="monospace", color=TEXT, fontweight="bold")
    ax.legend(fontsize=10, prop={"family": "monospace"})
    ax.tick_params(colors=MUTED, labelsize=10)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    fig.tight_layout()
    fig.savefig(out_path, dpi=200, facecolor=BG, bbox_inches="tight")
    plt.close(fig)


def plot_scurve(finishes: np.ndarray, deterministic: float, out_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(12, 6), facecolor=BG)
    ax.set_facecolor(BG)
    sorted_f = np.sort(finishes)
    cdf = np.arange(1, len(sorted_f) + 1) / len(sorted_f) * 100
    ax.plot(sorted_f, cdf, color=ACCENT, linewidth=2)

    p50, p80, p90 = (np.percentile(finishes, p) for p in [50, 80, 90])
    for pval, pname, color in [(p50, "P50", "#F59E0B"), (p80, "P80", "#F97316"), (p90, "P90", CRIT)]:
        ax.axvline(pval, color=color, linewidth=1.5, linestyle="--", alpha=0.8)
        ax.axhline(float(pname[1:]), color=color, linewidth=0.5, linestyle=":", alpha=0.5)
        ax.plot(pval, float(pname[1:]), "o", color=color, markersize=8)
        ax.annotate(f"{pname}: {pval:.0f}d", xy=(pval, float(pname[1:])),
                    xytext=(15, 5), textcoords="offset points",
                    fontsize=10, fontfamily="monospace", color=color, fontweight="bold")

    ax.axvline(deterministic, color=GREEN, linewidth=2, linestyle="--", alpha=0.8)
    det_pct = np.searchsorted(sorted_f, deterministic) / len(sorted_f) * 100
    ax.annotate(f"Deterministic: {deterministic:.0f}d\n({det_pct:.0f}% confidence)",
                xy=(deterministic, det_pct), xytext=(-15, 15), textcoords="offset points",
                fontsize=10, fontfamily="monospace", color=GREEN, fontweight="bold", ha="right")

    ax.set_xlabel("Project Duration (Working Days)", fontsize=11, fontfamily="monospace", color=TEXT)
    ax.set_ylabel("Cumulative Probability (%)", fontsize=11, fontfamily="monospace", color=TEXT)
    ax.set_title("S-Curve (CDF)\n"
                 "Probability of completing on or before a given duration",
                 fontsize=13, fontfamily="monospace", color=TEXT, fontweight="bold")
    ax.set_ylim(0, 105)
    ax.tick_params(colors=MUTED, labelsize=10)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.grid(axis="both", linestyle=":", alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=200, facecolor=BG, bbox_inches="tight")
    plt.close(fig)


def _spearman(x: np.ndarray, y: np.ndarray) -> float:
    rx = np.empty_like(x)
    ry = np.empty_like(y)
    rx[x.argsort()] = np.arange(len(x), dtype=float)
    ry[y.argsort()] = np.arange(len(y), dtype=float)
    d = rx - ry
    n = len(x)
    return 1 - 6 * np.sum(d ** 2) / (n * (n ** 2 - 1))


def plot_tornado(
    activities: dict[str, Activity], csv_order: list[str],
    project_finishes: np.ndarray, activity_durations: dict[str, np.ndarray],
    out_path: Path, top_n: int = 15,
) -> list[tuple[str, str, float, str]]:
    """Returns full correlation list (for writing to Excel) and saves chart."""
    correlations: list[tuple[str, str, float, str]] = []
    for aid in csv_order:
        act = activities[aid]
        corr = _spearman(activity_durations[aid], project_finishes)
        if not np.isnan(corr):
            correlations.append((aid, act.activity_name, corr, act.wbs))

    sorted_c = sorted(correlations, key=lambda x: abs(x[2]), reverse=True)
    top = sorted_c[:top_n]
    top.reverse()

    fig, ax = plt.subplots(figsize=(12, 8), facecolor=BG)
    ax.set_facecolor(BG)
    labels = [f"{aid}  {name}" for aid, name, _, _ in top]
    values = [c for _, _, c, _ in top]
    colors = [wbs_color(wbs) for _, _, _, wbs in top]

    ax.barh(range(len(top)), values, color=colors, edgecolor="#1f1f1f", linewidth=0.4, height=0.7)
    for i, (_, _, corr, _) in enumerate(top):
        ax.text(corr + 0.01 if corr >= 0 else corr - 0.01, i,
                f"{corr:.3f}", va="center", ha="left" if corr >= 0 else "right",
                fontsize=9, fontfamily="monospace", color=TEXT)

    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(labels, fontsize=9, fontfamily="monospace")
    ax.set_xlabel("Spearman Rank Correlation with Project Finish", fontsize=11,
                  fontfamily="monospace", color=TEXT)
    ax.set_title(f"Schedule Sensitivity (Top {top_n} Activities)\n"
                 "Spearman rank correlation of sampled duration vs. project completion",
                 fontsize=13, fontfamily="monospace", color=TEXT, fontweight="bold")
    ax.axvline(0, color=MUTED, linewidth=0.5)
    ax.tick_params(colors=MUTED, labelsize=10)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    fig.tight_layout()
    fig.savefig(out_path, dpi=200, facecolor=BG, bbox_inches="tight")
    plt.close(fig)

    return correlations


# ── Main ─────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Monte Carlo Schedule Risk Simulation (OPRA-compatible)")
    parser.add_argument("--schedule", type=Path, required=True,
                        help="Path to schedule CSV (activity network)")
    parser.add_argument("--risk-params", type=Path, default=None,
                        help="Path to risk parameters CSV (required with --init)")
    parser.add_argument("--workbook", type=Path, default=None,
                        help="Path to Excel workbook (created by --init, read on subsequent runs)")
    parser.add_argument("--output", type=Path, default=OUT,
                        help="Output directory for charts (default: outputs/)")
    parser.add_argument("--init", action="store_true",
                        help="Create the Excel workbook from CSVs")
    parser.add_argument("-n", "--iterations", type=int, default=DEFAULT_N,
                        help=f"Number of iterations (default: {DEFAULT_N:,})")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED,
                        help=f"Random seed (default: {DEFAULT_SEED})")
    args = parser.parse_args()

    out_dir = args.output
    out_dir.mkdir(exist_ok=True, parents=True)
    n_iter = args.iterations
    seed = args.seed

    # Default workbook path if not specified
    workbook_path = args.workbook or args.schedule.parent / (args.schedule.stem + "-monte-carlo.xlsx")

    activities, csv_order = load_schedule(args.schedule)

    # Pre-populate WBS color cache in schedule order for consistent coloring
    for aid in csv_order:
        wbs_color(activities[aid].wbs)

    if args.init:
        if args.risk_params is None:
            parser.error("--risk-params is required with --init")
        _create_workbook_from_args(activities, csv_order, args.risk_params, workbook_path)
        print(f"\nWorkbook ready for review: {workbook_path}")
        print("Edit the 'Duration Uncertainty' sheet,")
        print("then re-run without --init to simulate.")
        return

    if not workbook_path.exists():
        parser.error(f"Workbook not found: {workbook_path}\nRun with --init first to create it.")

    # Load assumptions from workbook
    risk_params = load_from_workbook(workbook_path)
    print(f"Loaded assumptions from: {workbook_path.name}")
    print(f"  {len(risk_params)} activity distributions")

    # Deterministic baseline
    topo = topological_order(activities, csv_order)
    det_durations = {aid: float(activities[aid].duration_days) for aid in csv_order}
    det_finish = forward_pass(topo, activities, det_durations)
    deterministic = max(det_finish.values())

    print(f"\nDeterministic project finish: {deterministic:.0f} working days")
    print(f"Running {n_iter:,} Monte Carlo iterations (seed={seed})...")

    # Simulate
    project_finishes, activity_durations = run_simulation(
        activities, csv_order, risk_params, n_iter, seed
    )

    # Stats
    p50, p80, p90 = (np.percentile(project_finishes, p) for p in [50, 80, 90])
    det_confidence = np.searchsorted(np.sort(project_finishes), deterministic) / n_iter * 100

    print(f"\n{'─' * 50}")
    print(f"  RESULTS ({n_iter:,} iterations)")
    print(f"{'─' * 50}")
    print(f"  Deterministic:  {deterministic:>7.0f}d  ({det_confidence:.0f}% confidence)")
    print(f"  Mean:           {np.mean(project_finishes):>7.0f}d")
    print(f"  Std Dev:        {np.std(project_finishes):>7.1f}d")
    print(f"  Min:            {np.min(project_finishes):>7.0f}d")
    print(f"  Max:            {np.max(project_finishes):>7.0f}d")
    print(f"  P50:            {p50:>7.0f}d")
    print(f"  P80:            {p80:>7.0f}d")
    print(f"  P90:            {p90:>7.0f}d")
    print(f"{'─' * 50}")

    # Charts
    plot_histogram(project_finishes, deterministic, out_dir / "monte-carlo-histogram.png")
    print("  → histogram saved")
    plot_scurve(project_finishes, deterministic, out_dir / "monte-carlo-scurve.png")
    print("  → S-curve saved")
    correlations = plot_tornado(activities, csv_order, project_finishes, activity_durations,
                                out_dir / "monte-carlo-tornado.png")
    print("  → tornado chart saved")

    # Write results back to workbook
    write_results_to_workbook(workbook_path, deterministic, project_finishes, correlations, seed)
    print("  → Results & Sensitivity sheets updated in workbook")

    print(f"\nAll outputs in: {out_dir}")
    print(f"Workbook:       {workbook_path}")


def _create_workbook_from_args(
    activities: dict[str, Activity], csv_order: list[str],
    risk_params_csv: Path, workbook_path: Path,
) -> None:
    """Build the master workbook from a schedule and risk params CSV."""
    rp: dict[str, dict] = {}
    with risk_params_csv.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rp[row["activity_id"]] = row

    wb = Workbook()

    # ── Duration Uncertainty sheet ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Duration Uncertainty"
    ws1.sheet_properties.tabColor = "3B82F6"

    headers1 = [
        "Activity ID", "Activity Name", "WBS", "WBS Name",
        "Baseline\n(days)", "Optimistic\nFactor", "Most Likely\nFactor",
        "Pessimistic\nFactor", "Optimistic\n(days)", "Most Likely\n(days)",
        "Pessimistic\n(days)", "Assumption Notes",
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header_row(ws1, 1, len(headers1))
    ws1.row_dimensions[1].height = 36

    for i, aid in enumerate(csv_order, 2):
        act = activities[aid]
        params = rp.get(aid, {})
        opt_f = float(params.get("optimistic_factor", 0.85))
        ml_f = float(params.get("most_likely_factor", 1.0))
        pes_f = float(params.get("pessimistic_factor", 1.4))
        notes = params.get("notes", "")

        ws1.cell(row=i, column=1, value=aid)
        ws1.cell(row=i, column=2, value=act.activity_name)
        ws1.cell(row=i, column=3, value=act.wbs)
        ws1.cell(row=i, column=4, value=act.wbs_name)
        ws1.cell(row=i, column=5, value=act.duration_days)
        ws1.cell(row=i, column=6, value=opt_f)
        ws1.cell(row=i, column=7, value=ml_f)
        ws1.cell(row=i, column=8, value=pes_f)
        ws1.cell(row=i, column=9).value = f"=E{i}*F{i}"
        ws1.cell(row=i, column=9).number_format = "0.0"
        ws1.cell(row=i, column=10).value = f"=E{i}*G{i}"
        ws1.cell(row=i, column=10).number_format = "0.0"
        ws1.cell(row=i, column=11).value = f"=E{i}*H{i}"
        ws1.cell(row=i, column=11).number_format = "0.0"
        ws1.cell(row=i, column=12, value=notes)

    end_row1 = 1 + len(csv_order)
    style_data_rows(ws1, 2, end_row1, len(headers1))

    for r in range(2, end_row1 + 1):
        for c in [6, 7, 8]:
            ws1.cell(row=r, column=c).number_format = "0.00"

    widths1 = [12, 38, 6, 28, 10, 12, 12, 12, 12, 12, 12, 48]
    for c, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(c)].width = w

    ws1.freeze_panes = "A2"

    # ── Placeholder sheets ───────────────────────────────────────────
    ws3 = wb.create_sheet("Results")
    ws3.sheet_properties.tabColor = "16A34A"
    ws3.cell(row=1, column=1, value="Run the simulation to populate this sheet.")
    ws3.cell(row=1, column=1).font = Font(italic=True, color="6B6B6B")

    ws4 = wb.create_sheet("Sensitivity")
    ws4.sheet_properties.tabColor = "F4A261"
    ws4.cell(row=1, column=1, value="Run the simulation to populate this sheet.")
    ws4.cell(row=1, column=1).font = Font(italic=True, color="6B6B6B")

    wb.save(workbook_path)
    print(f"  → workbook created: {workbook_path.name}")


if __name__ == "__main__":
    main()
